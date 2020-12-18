# Program that takes two integers and a filename string as command line arguments.
# Example command: py blankRowInserter.py 3 2 myProduce.xlsx
# With the above command line the program should start at row 3(n) and insert 2(m) blank rows into the excel spreadsheet.

import sys, openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# Interprets command line.
n = int(sys.argv[1])
m = int(sys.argv[2])
wb = openpyxl.load_workbook(sys.argv[3])
# Inserts blank rows into the spreadsheet.
sheet = wb.active
for i in range(n, n+m):
    sheet.insert_rows(n)
# Saves into a new spreadsheet.
wb.save('blankRows.xlsx')
print('Done.')


# # THIS CODE IS TO DELETE CELLS INSTEAD OF INSERTING THEM.
# sheet = wb.active
# for i in range(1, sheet.max_column + 1):
#     sheet[get_column_letter(i) + str(n)] = None
#     for x in range(n, n+m):
#         sheet[get_column_letter(i) + str(x)] = None